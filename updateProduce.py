import openpyxl

#Step-1 : Set up a data structure with the updated information
wb=openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

#The produce types and their updated prices
PRICE_UPDATES = { 'Garlic': 3.07,
                  'Celery': 1.19,
                  'Lemon': 1.27 
                 }

#Step 2 : Check all rows and update incorrect prices
for rowNum in range(2,sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')